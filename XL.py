'''
Created on 12 мая 2019 г.

@author: alex
'''

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
from config import *
from services import *

#Определяем путь к срипту и из папки уровня выше будем читать штатки
pathOfFile=os.getcwd()
path=os.path.split(pathOfFile)[0]
sourcePath=os.path.join(path, SOURCEDIR)
resultPath=os.path.join(path, RESULTDIR)
print (sourcePath)
print (resultPath)
#Перечень МРФ, которые есть в штатке
MRF={}

#Более новая штатка
WBNewest = load_workbook(os.path.join(sourcePath, "1.xlsx"), data_only=True)
print ("Загружена последняя штатная книга")
#Более старая штатка
WBOlder = load_workbook(os.path.join(sourcePath, "2.xlsx"), data_only=True)
print ("Загружена предыдущая штатная книга")

WSNewest = WBNewest.worksheets[0]
WSOlder = WBOlder.worksheets[0]


#Находим номера столбцов с табельными номерами
tabNewestNumber = getColumnID(WSNewest, columnName=EMPLOYERID)
tabOlderNumber = getColumnID(WSOlder, columnName=EMPLOYERID)
#Находим столбец с годовым ФОТ
tabOlderFOT=getColumnID(WSOlder, columnName=FOT)

#Получаем кортежи с табельными номерами. Устанавливаем min_row=2, что бы не брать заголовок таблицы.
rNew=tuple(WSNewest.iter_cols(min_col=tabNewestNumber, max_col=tabNewestNumber,min_row=2))
rOld=tuple(WSOlder.iter_cols(min_col=tabOlderNumber, max_col=tabOlderNumber, min_row=2))

#Увеличили на единицу, потому что не берем заголовок таблицы
lenRNew=len(rNew[0])+1

for i in rNew[0]:
    print (f"Обработана строка {i.row} из {lenRNew}")
    for y in rOld[0]:
        #Если Табельный номер из новой штатки без учета количества назначений присутствует в старой штатке, то запоминаем его для последюующего вывода.
        if (i.value!="") and (str(i.value).split("-")[0] == str(y.value).split("-")[0]):
            #Определяем МРФ. Если такого еще не было, то создаем новую книгу эксель и запоминаем ее в списке
            MRFName = WSNewest.cell(row=i.row, column=getColumnID(WSNewest, columnName="1")).value
          
            if MRFName not in MRF.keys():
                MRF[MRFName] = Workbook()
                MRF[MRFName].create_sheet("Лист 1", 0)               
                copyTableHeader(MRF[MRFName].worksheets[0], WSNewest)               
                print (MRF)

            #Добавляем в выходну книгу новую строку и копируем в нее запись из новой книги, значение ФОТ по человеку из предыдущей книги и формулу
            #для расчета отношения нового ФОТ к старому
            maxRow=MRF[MRFName].worksheets[0].max_row+1
            MRF[MRFName].worksheets[0].insert_rows(maxRow)           
                
            for w in WSNewest.iter_rows(min_row=i.row, max_row=i.row):
                for z in range(1, len(w)+1):
                    copyCell(MRF[MRFName].worksheets[0].cell(row=maxRow, column=z), WSNewest.cell(row=i.row, column=z))
                
                copyCell(MRF[MRFName].worksheets[0].cell(row=maxRow, column=len(w)+1), WSOlder.cell(row=y.row, column=tabOlderFOT))
                MRF[MRFName].worksheets[0].cell(row=maxRow, column=len(w)+2).value=f"={get_column_letter(len(w))}{maxRow}/{(get_column_letter(len(w)+1))}{maxRow}-1"
                MRF[MRFName].worksheets[0].cell(row=maxRow, column=len(w)+2).number_format = "0.00%"

#Сохраняем все полученные книги
for i in MRF:
    MRF[i].save(os.path.join(resultPath, f"{i}.xlsx"))
    MRF[i].close()
      
WBNewest.close()
WBOlder.close()