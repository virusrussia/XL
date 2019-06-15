'''
Created on 14 июн. 2019 г.

@author: alex
'''
from config import *
import openpyxl
from copy import copy
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.copier import WorksheetCopy

def getColumnID (WB, columnName=EMPLOYERID):
   "Ищем номер столбца с заданным именем"
   
   for i in range(1, WB.max_column+1):
       if str(WB.cell(row=1, column=i).value) == columnName:
           return i

 
def copyCell(cellResult, cellSource):
    "Полностью копируем ячейку"
    cellResult.data_type = cellSource.data_type
    cellResult.value = cellSource.value
    
    if cellSource.has_style:
        cellResult.font = copy(cellSource.font)
        cellResult.border = copy(cellSource.border)
        cellResult.fill = copy(cellSource.fill)
        cellResult.number_format = copy(cellSource.number_format)
        cellResult.protection = copy(cellSource.protection)
        cellResult.alignment = copy(cellSource.alignment)

       

def copyTableHeader (WSResult, WSSource):
    "Копируем заголовок таблицы и добавляем еще два поля в конце"
    for i in range(1, WSSource.max_column+1):
           copyCell(WSResult.cell(row=1, column=i), WSSource.cell(row=1, column=i))
 
    m=WSSource.max_column
    copyCell(WSResult.cell(row=1, column=m+1),WSSource.cell(row=1, column=m))
    WSResult.cell(row=1, column=m+1).value="ФОТ в предыдущем срезе"

    copyCell(WSResult.cell(row=1, column=m+2),WSSource.cell(row=1, column=m))   
    WSResult.cell(row=1, column=m+2).value="Δ"


def copyRow (rowResult, rowSource):
       for i in range(1, rowSource.max_column+1):
           copyCell(rowResult.cell(row=1, column=i), rowSource.cell(row=1, column=i)) 
    
    
        
        
